# -*- coding: utf-8 -*-
"""Excel 读写辅助函数（最新主链路）。

提供 CSV/TSV 多编码读取、openpyxl 工作簿操作、列头提取与重复检测、
公式单元格统计，以及通过 LibreOffice 宏进行公式重算的完整链路。
"""
from __future__ import annotations

import csv
import os
import platform
import shutil
import socket
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from common import json_friendly


def try_read_csv(path: Path, delimiter: str, count_rows: bool = False) -> tuple[list[str], int, str, str]:
    """尝试按多种编码读取 CSV/TSV，返回表头和行数。"""
    last_error = ""
    for encoding in ("utf-8-sig", "gb18030", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                reader = csv.reader(f, delimiter=delimiter)
                first = next(reader, None)
                headers = [str(v).strip() for v in first] if first else []
                row_count = sum(1 for _ in reader) if count_rows else 0
                return headers, row_count, encoding, ""
        except UnicodeDecodeError as e:
            last_error = str(e)
    return [], 0, "", last_error


def open_workbook(path: Path, data_only: bool = True):
    """按只读模式打开工作簿，不做兼容回退。"""
    from openpyxl import load_workbook

    return load_workbook(path, read_only=True, data_only=data_only)


def extract_headers(first_row: tuple) -> list[str]:
    """从首行提取列名。"""
    return ["" if v is None else str(json_friendly(v)).strip() for v in first_row]


def detect_duplicate_headers(headers: list[str]) -> str:
    """检查重复列名并返回告警文本。"""
    seen: dict[str, int] = {}
    for h in headers:
        if h:
            seen[h] = seen.get(h, 0) + 1
    dups = {k: v for k, v in seen.items() if v > 1}
    if dups:
        parts = [f"'{k}'({v}次)" for k, v in sorted(dups.items())]
        return f"检测到重复列名：{', '.join(parts)}。重复列可能导致数据覆盖"
    return ""


def count_formula_cells(path: Path) -> int:
    """统计工作簿中的公式单元格数量。"""
    wb = open_workbook(path, data_only=False)
    count = 0
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str) and value.startswith("="):
                        count += 1
    finally:
        wb.close()
    return count


EXCEL_ERROR_TOKENS = (
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)
MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""
LINUX_SHIM_SO = Path(tempfile.gettempdir()) / "ecv_soffice_socket_shim.so"
LINUX_SHIM_C = Path(tempfile.gettempdir()) / "ecv_soffice_socket_shim.c"
LINUX_SHIM_SOURCE = r"""
#define _GNU_SOURCE
#include <dlfcn.h>
#include <errno.h>
#include <signal.h>
#include <stdio.h>
#include <stdlib.h>
#include <sys/socket.h>
#include <unistd.h>

static int (*real_socket)(int, int, int);
static int (*real_socketpair)(int, int, int, int[2]);
static int (*real_listen)(int, int);
static int (*real_accept)(int, struct sockaddr *, socklen_t *);
static int (*real_close)(int);
static int (*real_read)(int, void *, size_t);

/* Per-FD bookkeeping (FDs >= 1024 are passed through unshimmed). */
static int is_shimmed[1024];
static int peer_of[1024];
static int wake_r[1024];            /* accept() blocks reading this */
static int wake_w[1024];            /* close()  writes to this      */
static int listener_fd = -1;        /* FD that received listen()    */

__attribute__((constructor))
static void init(void) {
    real_socket     = dlsym(RTLD_NEXT, "socket");
    real_socketpair = dlsym(RTLD_NEXT, "socketpair");
    real_listen     = dlsym(RTLD_NEXT, "listen");
    real_accept     = dlsym(RTLD_NEXT, "accept");
    real_close      = dlsym(RTLD_NEXT, "close");
    real_read       = dlsym(RTLD_NEXT, "read");
    for (int i = 0; i < 1024; i++) {
        peer_of[i] = -1;
        wake_r[i]  = -1;
        wake_w[i]  = -1;
    }
}

/* ---- socket ---------------------------------------------------------- */
int socket(int domain, int type, int protocol) {
    if (domain == AF_UNIX) {
        int fd = real_socket(domain, type, protocol);
        if (fd >= 0) return fd;
        /* socket(AF_UNIX) blocked – fall back to socketpair(). */
        int sv[2];
        if (real_socketpair(domain, type, protocol, sv) == 0) {
            if (sv[0] >= 0 && sv[0] < 1024) {
                is_shimmed[sv[0]] = 1;
                peer_of[sv[0]]    = sv[1];
                int wp[2];
                if (pipe(wp) == 0) {
                    wake_r[sv[0]] = wp[0];
                    wake_w[sv[0]] = wp[1];
                }
            }
            return sv[0];
        }
        errno = EPERM;
        return -1;
    }
    return real_socket(domain, type, protocol);
}

/* ---- listen ---------------------------------------------------------- */
int listen(int sockfd, int backlog) {
    if (sockfd >= 0 && sockfd < 1024 && is_shimmed[sockfd]) {
        listener_fd = sockfd;
        return 0;
    }
    return real_listen(sockfd, backlog);
}

/* ---- accept ---------------------------------------------------------- */
int accept(int sockfd, struct sockaddr *addr, socklen_t *addrlen) {
    if (sockfd >= 0 && sockfd < 1024 && is_shimmed[sockfd]) {
        /* Block until close() writes to the wake pipe. */
        if (wake_r[sockfd] >= 0) {
            char buf;
            real_read(wake_r[sockfd], &buf, 1);
        }
        errno = ECONNABORTED;
        return -1;
    }
    return real_accept(sockfd, addr, addrlen);
}

/* ---- close ----------------------------------------------------------- */
int close(int fd) {
    if (fd >= 0 && fd < 1024 && is_shimmed[fd]) {
        int was_listener = (fd == listener_fd);
        is_shimmed[fd] = 0;

        if (wake_w[fd] >= 0) {              /* unblock accept() */
            char c = 0;
            write(wake_w[fd], &c, 1);
            real_close(wake_w[fd]);
            wake_w[fd] = -1;
        }
        if (wake_r[fd] >= 0) { real_close(wake_r[fd]); wake_r[fd]  = -1; }
        if (peer_of[fd] >= 0) { real_close(peer_of[fd]); peer_of[fd] = -1; }

        if (was_listener)
            _exit(0);                        /* conversion done – exit */
    }
    return real_close(fd);
}
"""


def _needs_linux_socket_shim() -> bool:
    """检测当前 Linux 环境是否需要 AF_UNIX socket shim。"""
    if platform.system() != "Linux":
        return False
    try:
        sock = socket.socket(socket.AF_UNIX, socket.SOCK_STREAM)
        sock.close()
        return False
    except OSError:
        return True


def _ensure_linux_socket_shim() -> tuple[str | None, str]:
    """确保 socket shim 共享库已编译就绪，返回 (路径, 错误)。"""
    if LINUX_SHIM_SO.exists():
        return str(LINUX_SHIM_SO), ""

    try:
        LINUX_SHIM_C.write_text(LINUX_SHIM_SOURCE, encoding="utf-8")
        result = subprocess.run(
            ["gcc", "-shared", "-fPIC", "-o", str(LINUX_SHIM_SO), str(LINUX_SHIM_C), "-ldl"],
            capture_output=True,
            text=True,
            check=False,
            timeout=20,
        )
        if result.returncode != 0:
            err = (result.stderr or result.stdout or "").strip()
            return None, err or f"gcc 编译失败，退出码={result.returncode}"
        return str(LINUX_SHIM_SO), ""
    except Exception as exc:  # noqa: BLE001
        return None, str(exc)
    finally:
        try:
            if LINUX_SHIM_C.exists():
                LINUX_SHIM_C.unlink()
        except Exception:  # noqa: BLE001
            pass


def _build_soffice_env() -> tuple[dict[str, str], list[str]]:
    """构建 LibreOffice 子进程环境变量，返回 (env, notes)。"""
    env = os.environ.copy()
    notes: list[str] = []
    env["SAL_USE_VCLPLUGIN"] = "svp"
    notes.append("启用 SAL_USE_VCLPLUGIN=svp")

    if platform.system() != "Linux":
        return env, notes

    if not _needs_linux_socket_shim():
        notes.append("Linux 环境 AF_UNIX 可用，无需 socket shim")
        return env, notes

    shim_so, err = _ensure_linux_socket_shim()
    if shim_so:
        existing = str(env.get("LD_PRELOAD", "")).strip()
        env["LD_PRELOAD"] = f"{shim_so}:{existing}" if existing else shim_so
        notes.append(f"Linux 环境启用 socket shim：{shim_so}")
    else:
        notes.append(f"Linux 环境无法启用 socket shim，已降级继续：{err}")
    return env, notes


def _resolve_soffice_binary() -> str | None:
    """查找 soffice 可执行文件路径，未找到时返回 None。"""
    bin_path = shutil.which("soffice")
    if bin_path:
        return bin_path
    if platform.system() != "Windows":
        return None

    candidates = [
        Path(os.environ.get("ProgramFiles", "")) / "LibreOffice" / "program" / "soffice.exe",
        Path(os.environ.get("ProgramFiles(x86)", "")) / "LibreOffice" / "program" / "soffice.exe",
    ]
    for path in candidates:
        if path.exists():
            return str(path)
    return None


def _macro_dir() -> Path:
    """返回当前平台下 LibreOffice 用户宏目录。"""
    system_name = platform.system()
    if system_name == "Windows":
        appdata = os.environ.get("APPDATA")
        if appdata:
            return Path(appdata) / "LibreOffice" / "4" / "user" / "basic" / "Standard"
        return Path.home() / "AppData" / "Roaming" / "LibreOffice" / "4" / "user" / "basic" / "Standard"
    if system_name == "Darwin":
        return Path.home() / "Library" / "Application Support" / "LibreOffice" / "4" / "user" / "basic" / "Standard"
    return Path.home() / ".config" / "libreoffice" / "4" / "user" / "basic" / "Standard"


def _ensure_libreoffice_macro(soffice_bin: str, soffice_env: dict[str, str]) -> tuple[bool, str]:
    """确保 RecalculateAndSave 宏已写入用户目录，返回 (成功, 错误)。"""
    macro_dir = _macro_dir()
    macro_file = macro_dir / MACRO_FILENAME

    if macro_file.exists():
        try:
            content = macro_file.read_text(encoding="utf-8", errors="ignore")
            if "RecalculateAndSave" in content:
                return True, ""
        except Exception:  # noqa: BLE001
            pass

    try:
        subprocess.run(
            [soffice_bin, "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=20,
            check=False,
            env=soffice_env,
        )
    except Exception:  # noqa: BLE001
        pass

    try:
        macro_dir.mkdir(parents=True, exist_ok=True)
        macro_file.write_text(RECALCULATE_MACRO, encoding="utf-8")
        return True, ""
    except Exception as exc:  # noqa: BLE001
        return False, str(exc)


def _run_recalc_macro(
    path: Path,
    soffice_bin: str,
    timeout_seconds: int,
    soffice_env: dict[str, str],
) -> dict[str, Any]:
    """调用 LibreOffice 宏执行公式重算，返回结果字典。"""
    macro_uri = "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application"
    cmd = [
        soffice_bin,
        "--headless",
        "--norestore",
        macro_uri,
        str(path.resolve()),
    ]
    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=max(5, int(timeout_seconds or 30)),
            check=False,
            env=soffice_env,
        )
    except subprocess.TimeoutExpired:
        return {"error": f"公式重算超时（>{timeout_seconds}秒）"}
    except Exception as exc:  # noqa: BLE001
        return {"error": f"调用 LibreOffice 失败：{exc}"}

    if result.returncode != 0:
        raw_error = (result.stderr or result.stdout or "").strip()
        if "Module1" in raw_error or "RecalculateAndSave" in raw_error:
            return {"error": "LibreOffice 宏未正确加载，请检查用户目录写权限或 LibreOffice 安装状态"}
        return {"error": raw_error or f"LibreOffice 公式重算失败，退出码={result.returncode}"}
    return {}


def recalc_excel_with_libreoffice(path: Path, timeout_seconds: int = 30) -> dict[str, Any]:
    """按内置 LibreOffice 主链路重算公式并返回结果摘要。"""
    runtime_notes: list[str] = []

    if not path.exists():
        return {"error": f"文件不存在：{path}", "runtime_notes": runtime_notes}

    soffice_env, env_notes = _build_soffice_env()
    runtime_notes.extend(env_notes)

    soffice_bin = _resolve_soffice_binary()
    if soffice_bin is None:
        return {"error": "未找到 soffice，可安装 LibreOffice 或将 soffice 加入 PATH", "runtime_notes": runtime_notes}

    ok, err = _ensure_libreoffice_macro(soffice_bin, soffice_env)
    if not ok:
        return {"error": f"初始化 LibreOffice 宏失败：{err}", "runtime_notes": runtime_notes}

    run_result = _run_recalc_macro(path, soffice_bin, timeout_seconds, soffice_env)
    if run_result.get("error"):
        run_result["runtime_notes"] = runtime_notes
        return run_result

    try:
        wb_data = open_workbook(path, data_only=True)
        error_details = {token: [] for token in EXCEL_ERROR_TOKENS}
        total_errors = 0
        try:
            for sheet_name in wb_data.sheetnames:
                ws = wb_data[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if not isinstance(cell.value, str):
                            continue
                        val = cell.value.upper()
                        for token in EXCEL_ERROR_TOKENS:
                            if token in val:
                                error_details[token].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        finally:
            wb_data.close()
    except Exception as exc:  # noqa: BLE001
        return {"error": f"重算完成后读取工作簿失败：{exc}", "runtime_notes": runtime_notes}

    summary: dict[str, Any] = {}
    for token, locations in error_details.items():
        if locations:
            summary[token] = {
                "count": len(locations),
                "locations": locations[:20],
            }

    try:
        formula_count = count_formula_cells(path)
    except Exception:  # noqa: BLE001
        formula_count = 0

    return {
        "status": "success" if total_errors == 0 else "errors_found",
        "total_errors": total_errors,
        "error_summary": summary,
        "total_formulas": formula_count,
        "runtime_notes": runtime_notes,
    }
