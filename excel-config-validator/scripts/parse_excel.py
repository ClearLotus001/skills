"""Excel/CSV 解析引擎。

用于读取输入文件并生成流式行存储（JSONL），供后续校验阶段复用。
"""
from __future__ import annotations

import argparse
import csv
import fnmatch
import hashlib
import json
import sys
import warnings
from pathlib import Path
from typing import Any
from zipfile import ZipFile

# 确保以任意工作目录执行脚本时都能正确导入同级模块
sys.path.insert(0, str(Path(__file__).resolve().parent))

from common import atomic_write_json, dataset_configs, file_sha256, json_friendly, normalize_path_text, utc_now_iso
from excel_io import (
    count_formula_cells,
    detect_duplicate_headers,
    extract_headers,
    open_workbook,
    recalc_excel_with_libreoffice,
    try_read_csv,
)
from xlsx_package_check import check_xlsx_package


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv", ".tsv"}
IGNORED_ARTIFACT_NAMES = {
    "issues.csv",
    "issues_raw.csv",
}
TEXT_SHEET_NAME = "_csv_"


def text_format_label(path: Path) -> str:
    return "TSV" if path.suffix.lower() == ".tsv" else "CSV"


def text_delimiter(path: Path) -> str:
    return "\t" if path.suffix.lower() == ".tsv" else ","


def discover_input_files(inputs: Path) -> list[Path]:
    if inputs.is_file():
        ext = inputs.suffix.lower()
        if inputs.name.startswith("~$"):
            return []
        if inputs.name.lower() in IGNORED_ARTIFACT_NAMES:
            return []
        return [inputs] if ext in SUPPORTED_EXTENSIONS else []
    files: list[Path] = []
    for path in inputs.rglob("*"):
        if (
            path.is_file()
            and path.suffix.lower() in SUPPORTED_EXTENSIONS
            and not path.name.startswith("~$")
            and path.name.lower() not in IGNORED_ARTIFACT_NAMES
        ):
            files.append(path)
    return sorted(files)


def row_to_map(headers: list[str], row_values: list[Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    width = max(len(headers), len(row_values))
    for idx in range(width):
        key = headers[idx] if idx < len(headers) else ""
        key = str(key).strip()
        if not key:
            key = f"_col_{idx + 1}"
        out[key] = json_friendly(row_values[idx] if idx < len(row_values) else None)
    return out


# normalize_path_text 与 dataset_configs 由 common 模块导入


def build_projection_plan(compiled_rules_path: Path | None) -> list[dict[str, Any]]:
    if compiled_rules_path is None or not compiled_rules_path.exists():
        return []

    data = json.loads(compiled_rules_path.read_text(encoding="utf-8"))
    rules = data.get("rules", {})
    if not isinstance(rules, dict):
        return []

    datasets = dataset_configs(rules)
    ds_columns: dict[str, set[str]] = {k: set() for k in datasets.keys()}
    ds_project_all: set[str] = set()

    for key in ("schema_rules", "range_rules"):
        items = rules.get(key, [])
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            ds = str(item.get("dataset", "")).strip()
            col = str(item.get("column", "")).strip()
            if ds and ds in ds_columns and col:
                ds_columns[ds].add(col)

    relation_rules = rules.get("relation_rules", [])
    if isinstance(relation_rules, list):
        for item in relation_rules:
            if not isinstance(item, dict):
                continue
            source_ds = str(item.get("source_dataset") or "").strip()
            target_ds = str(item.get("target_dataset") or "").strip()
            source_key = str(item.get("source_key") or "").strip()
            target_key = str(item.get("target_key") or "").strip()
            if source_ds and source_ds in ds_columns and source_key:
                ds_columns[source_ds].add(source_key)
            if target_ds and target_ds in ds_columns and target_key:
                ds_columns[target_ds].add(target_key)

    row_rules = rules.get("row_rules", [])
    if isinstance(row_rules, list):
        for item in row_rules:
            if not isinstance(item, dict):
                continue
            ds = str(item.get("dataset", "")).strip()
            if ds and ds in ds_columns:
                ds_project_all.add(ds)
                col = str(item.get("column", "")).strip()
                if col:
                    ds_columns[ds].add(col)

    plan: list[dict[str, Any]] = []
    for ds, cfg in datasets.items():
        file_name = str(cfg.get("file", "")).strip()
        file_pattern = str(cfg.get("file_pattern", "")).strip()
        sheet = str(cfg.get("sheet", "")).strip()
        plan.append(
            {
                "dataset": ds,
                "file": file_name,
                "file_pattern": file_pattern,
                "sheet": sheet,
                "project_all": ds in ds_project_all,
                "columns": sorted(ds_columns.get(ds, set())),
            }
        )
    return plan


def projection_item_matches(
    item: dict[str, Any],
    *,
    file_name: str,
    file_path: str,
    sheet: str,
) -> bool:
    expected_file = normalize_path_text(str(item.get("file", "")))
    file_pattern = normalize_path_text(str(item.get("file_pattern", "")))
    expected_sheet = str(item.get("sheet", "")).strip()
    normalized_file_name = normalize_path_text(file_name)
    normalized_file_path = normalize_path_text(file_path)

    if expected_sheet and expected_sheet != sheet:
        return False
    if expected_file:
        return (
            normalized_file_name == expected_file
            or normalized_file_path == expected_file
            or normalized_file_path.endswith(f"/{expected_file}")
        )
    if file_pattern:
        return fnmatch.fnmatch(normalized_file_name, file_pattern) or fnmatch.fnmatch(normalized_file_path, file_pattern)
    return True


def projected_headers_for_sheet(
    plan: list[dict[str, Any]],
    *,
    file_name: str,
    file_path: str,
    sheet: str,
    headers: list[str],
) -> list[str]:
    if not headers:
        return []
    if not plan:
        return list(headers)

    matched = [
        item
        for item in plan
        if projection_item_matches(item, file_name=file_name, file_path=file_path, sheet=sheet)
    ]
    if not matched:
        return []
    if any(bool(item.get("project_all")) for item in matched):
        return list(headers)

    col_set: set[str] = set()
    for item in matched:
        for col in item.get("columns", []) or []:
            x = str(col).strip()
            if x:
                col_set.add(x)
    if not col_set:
        return list(headers)

    filtered = [h for h in headers if h in col_set]
    return filtered or list(headers)


def rows_store_path(out_dir: Path, file_path: Path, sheet_name: str) -> Path:
    key = hashlib.sha1(f"{file_path.as_posix()}::{sheet_name}".encode("utf-8")).hexdigest()[:16]
    safe_sheet = "".join(ch if ch.isalnum() else "_" for ch in sheet_name)[:48] or "sheet"
    store_dir = out_dir / "_row_store"
    store_dir.mkdir(parents=True, exist_ok=True)
    return store_dir / f"{key}_{safe_sheet}.jsonl"


def write_rows_store_stream(
    rows_iter: Any,
    projected_headers: list[str] | None,
    output_path: Path,
    chunk_size: int,
) -> int:
    keep_all = projected_headers is None
    selected_headers = projected_headers or []
    size = max(1, int(chunk_size or 1))
    count = 0
    buffer: list[str] = []
    with output_path.open("w", encoding="utf-8", newline="\n") as f:
        for row_no, values in rows_iter:
            if not isinstance(values, dict):
                values = {}
            if keep_all:
                new_values = values
            else:
                new_values = {k: values.get(k) for k in selected_headers if k in values}
            payload = {"row": row_no, "values": new_values}
            buffer.append(json.dumps(payload, ensure_ascii=False))
            count += 1
            if len(buffer) >= size:
                f.write("\n".join(buffer))
                f.write("\n")
                buffer.clear()
        if buffer:
            f.write("\n".join(buffer))
            f.write("\n")
    return count


def scan_xlsx_extlst_sheet_xml(path: Path) -> list[str]:
    out: list[str] = []
    try:
        with ZipFile(path) as zf:
            for name in zf.namelist():
                if not (name.startswith("xl/worksheets/sheet") and name.endswith(".xml")):
                    continue
                data = zf.read(name)
                if b"<extLst" in data:
                    out.append(name.rsplit("/", 1)[-1])
    except Exception:  # noqa: BLE001
        return []
    return out


def select_sheet_region(ws: Any) -> tuple[int, int, int, int, str]:
    """优先按 Excel 表格（ListObject）区域解析；无表格时回退整表区域。"""
    try:
        from openpyxl.utils.cell import range_boundaries  # type: ignore
    except Exception:  # noqa: BLE001
        range_boundaries = None  # type: ignore[assignment]

    regions: list[dict[str, Any]] = []
    if range_boundaries is not None:
        try:
            tables = getattr(ws, "tables", {}) or {}
            for table in tables.values():
                ref = str(getattr(table, "ref", "") or "").strip()
                if not ref:
                    continue
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(ref)
                except Exception:  # noqa: BLE001
                    continue
                if max_row < min_row or max_col < min_col:
                    continue
                area = (max_row - min_row + 1) * (max_col - min_col + 1)
                regions.append(
                    {
                        "name": str(getattr(table, "name", "") or "").strip(),
                        "min_col": min_col,
                        "min_row": min_row,
                        "max_col": max_col,
                        "max_row": max_row,
                        "area": area,
                    }
                )
        except Exception:  # noqa: BLE001
            regions = []

    if regions:
        regions.sort(key=lambda x: int(x.get("area", 0) or 0), reverse=True)
        chosen = regions[0]
        note = ""
        if len(regions) > 1:
            note = (
                f"检测到多个 Excel 表格（{len(regions)} 个），"
                f"已使用数据区域最大的表格 '{chosen.get('name') or 'unnamed'}' 作为解析范围"
            )
        return (
            int(chosen["min_row"]),
            int(chosen["max_row"]),
            int(chosen["min_col"]),
            int(chosen["max_col"]),
            note,
        )

    max_row = ws.max_row
    max_col = ws.max_column
    safe_max_row = int(max_row) if isinstance(max_row, int) and max_row > 0 else 1
    safe_max_col = int(max_col) if isinstance(max_col, int) and max_col > 0 else 1
    return (1, safe_max_row, 1, safe_max_col, "")


def load_sheet_region_map(path: Path, data_only: bool) -> dict[str, tuple[int, int, int, int, str]]:
    """使用非只读工作簿提取各工作表的最优解析区域。"""
    region_map: dict[str, tuple[int, int, int, int, str]] = {}
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:  # noqa: BLE001
        return region_map

    try:
        wb_meta = load_workbook(path, read_only=False, data_only=data_only)
    except Exception:  # noqa: BLE001
        return region_map

    try:
        for sheet_name in wb_meta.sheetnames:
            region_map[sheet_name] = select_sheet_region(wb_meta[sheet_name])
    finally:
        wb_meta.close()

    return region_map


def _scan_text_file(path: Path) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    parse_warnings: list[str] = []
    notes: list[str] = []
    label = text_format_label(path)
    delimiter = text_delimiter(path)

    headers, row_count_estimate, used_encoding, error = try_read_csv(path, delimiter, count_rows=True)

    if not used_encoding:
        parse_warnings.append(f"{label} 编码无法识别（已尝试 UTF-8/GB18030/UTF-16）" + (f"：{error}" if error else ""))
        return ([{"sheet": TEXT_SHEET_NAME, "headers": [], "row_count_estimate": 0}], parse_warnings, notes)

    if used_encoding != "utf-8-sig":
        notes.append(f"{label} 非 UTF-8 编码，已自动按 {used_encoding} 读取")
    if not headers:
        parse_warnings.append(f"{label} 为空或缺少表头行")

    dup_warning = detect_duplicate_headers(headers)
    if dup_warning:
        parse_warnings.append(f"[{path.name}/{TEXT_SHEET_NAME}] {dup_warning}")

    return (
        [{"sheet": TEXT_SHEET_NAME, "headers": headers, "row_count_estimate": row_count_estimate}],
        parse_warnings,
        notes,
    )


def _scan_xlsx_file(path: Path) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    sheets: list[dict[str, Any]] = []
    parse_warnings: list[str] = []
    parse_notes: list[str] = []

    try:
        wb = open_workbook(path, data_only=True)
    except ImportError:
        return sheets, ["缺少 openpyxl，无法解析 .xlsx/.xlsm 表头"], parse_notes
    except Exception as e:  # noqa: BLE001
        parse_warnings.append(f"openpyxl 无法打开文件：{e}")
        return sheets, parse_warnings, parse_notes

    unknown_extension_hits = 0
    region_map = load_sheet_region_map(path, data_only=True) if getattr(wb, "read_only", False) else {}
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                min_row, max_row, min_col, max_col, region_note = region_map.get(sheet_name) or select_sheet_region(ws)
                if region_note:
                    parse_notes.append(f"[{path.name}/{sheet_name}] {region_note}")
                first_row = next(
                    ws.iter_rows(
                        min_row=min_row,
                        max_row=min_row,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True,
                    ),
                    tuple(),
                )
                headers = extract_headers(first_row)
                dup_warning = detect_duplicate_headers(headers)
                if dup_warning:
                    parse_warnings.append(f"[{path.name}/{sheet_name}] {dup_warning}")
                sheets.append(
                    {
                        "sheet": sheet_name,
                        "headers": headers,
                        "row_count_estimate": max(0, int(max_row) - int(min_row)),
                    }
                )
        finally:
            wb.close()

    for w in caught:
        msg = str(w.message)
        if "Unknown extension is not supported and will be removed" in msg:
            unknown_extension_hits += 1
            continue
        parse_warnings.append(f"openpyxl 告警：{msg}")

    if unknown_extension_hits > 0:
        xml_hits = scan_xlsx_extlst_sheet_xml(path)
        xml_text = f"；涉及: {', '.join(xml_hits[:6])}" if xml_hits else ""
        parse_notes.append(
            "检测到 openpyxl 不支持的工作表扩展（extLst）"
            f"，命中 {unknown_extension_hits} 次。根因通常是 Excel 高级特性扩展节点（如部分条件格式/数据验证/切片器）"
            f"{xml_text}。该扩展会被 openpyxl 忽略，但常规单元格值读取仍可继续。"
        )

    return sheets, parse_warnings, parse_notes


def duplicate_file_groups(files: list[Path]) -> dict[str, list[Path]]:
    groups: dict[str, list[Path]] = {}
    for file_path in files:
        groups.setdefault(file_path.name, []).append(file_path)
    return {name: paths for name, paths in groups.items() if len(paths) > 1}


def build_scan_payload(inputs: Path) -> dict[str, Any]:
    files = discover_input_files(inputs)
    if not files:
        raise ValueError(f"在 {inputs} 未发现可支持的输入文件")

    results: list[dict[str, Any]] = []
    for file_path in files:
        ext = file_path.suffix.lower()
        if ext in {".csv", ".tsv"}:
            sheets, parse_warnings, parse_notes = _scan_text_file(file_path)
        elif ext in {".xlsx", ".xlsm"}:
            sheets, parse_warnings, parse_notes = _scan_xlsx_file(file_path)
        else:
            sheets = []
            parse_warnings = [f"不支持的文件后缀：{ext}"]
            parse_notes = []
        sheet_map = {
            str(item.get("sheet", "")): {
                "headers": [str(x) for x in (item.get("headers", []) or [])],
                "row_count_estimate": int(item.get("row_count_estimate", 0) or 0),
            }
            for item in sheets
        }
        file_info: dict[str, Any] = {
            "file": file_path.name,
            "type": file_path.suffix.lower().lstrip("."),
            "sheets": sheet_map,
        }
        if parse_warnings:
            file_info["warnings"] = parse_warnings
        if parse_notes:
            file_info["notes"] = parse_notes
        results.append(file_info)

    duplicates = duplicate_file_groups(files)
    payload: dict[str, Any] = {"files": results, "file_count": len(results)}
    if duplicates:
        payload["duplicate_file_names"] = {
            name: [str(path) for path in paths]
            for name, paths in sorted(duplicates.items())
        }
    return payload


def write_scan_payload(inputs: Path, out_dir: Path) -> Path:
    payload = build_scan_payload(inputs=inputs)
    scan_path = out_dir / "_scan.json"
    atomic_write_json(scan_path, payload)
    return scan_path


def _ingest_text_file(
    *,
    file_path: Path,
    out_dir: Path,
    projection_plan: list[dict[str, Any]],
    row_chunk_size: int,
) -> tuple[list[dict[str, Any]], list[str], list[str], int]:
    parse_warnings: list[str] = []
    notes: list[str] = []
    row_store_files = 0
    label = text_format_label(file_path)
    delimiter = text_delimiter(file_path)
    sheet_entries: list[dict[str, Any]] = []

    headers, _, used_encoding, error = try_read_csv(file_path, delimiter, count_rows=False)

    if not used_encoding:
        if not any(w.startswith(f"{label} 解析失败：") for w in parse_warnings):
            parse_warnings.append(
                f"{label} 编码无法识别（已尝试 UTF-8/GB18030/UTF-16）"
                + (f"：{error}" if error else "")
            )
        if not sheet_entries:
            sheet_entries.append(
                {
                    "sheet": TEXT_SHEET_NAME,
                    "headers": [],
                    "projected_headers": [],
                    "row_count_estimate": 0,
                    "rows_file": "",
                }
            )
        return sheet_entries, parse_warnings, notes, row_store_files

    # 重新打开文件以执行完整读取
    try:
        with file_path.open("r", encoding=used_encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            header_row = next(reader, None)
            headers = [str(v).strip() for v in header_row] if header_row is not None else []
            if not headers:
                parse_warnings.append(f"{label} 为空或缺少表头行")
            sheet_name = TEXT_SHEET_NAME
            dup_warning = detect_duplicate_headers(headers)
            if dup_warning:
                parse_warnings.append(f"[{file_path.name}/{sheet_name}] {dup_warning}")

            projected_headers = projected_headers_for_sheet(
                projection_plan,
                file_name=file_path.name,
                file_path=file_path.as_posix(),
                sheet=sheet_name,
                headers=headers,
            )

            store_path_text = ""
            if not (projection_plan and not projected_headers):
                store_path = rows_store_path(out_dir, file_path, sheet_name)
                row_count = write_rows_store_stream(
                    rows_iter=(
                        (row_idx, row_to_map(headers, list(row)))
                        for row_idx, row in enumerate(reader, start=2)
                    ),
                    projected_headers=projected_headers,
                    output_path=store_path,
                    chunk_size=row_chunk_size,
                )
                store_path_text = store_path.as_posix()
                row_store_files += 1
            else:
                row_count = sum(1 for _ in reader)

            sheet_entries.append(
                {
                    "sheet": sheet_name,
                    "headers": headers,
                    "projected_headers": projected_headers,
                    "row_count_estimate": row_count,
                    "rows_file": store_path_text,
                }
            )
    except Exception as exc:  # noqa: BLE001
        parse_warnings.append(f"{label} 解析失败：{exc}")

    if used_encoding != "utf-8-sig":
        notes.append(f"{label} 非 UTF-8 编码，已自动按 {used_encoding} 读取")

    return sheet_entries, parse_warnings, notes, row_store_files


def _ingest_xlsx_file(
    *,
    file_path: Path,
    out_dir: Path,
    projection_plan: list[dict[str, Any]],
    row_chunk_size: int,
    skip_xlsx_package_check: bool,
) -> tuple[list[dict[str, Any]], list[str], list[str], int]:
    sheet_entries: list[dict[str, Any]] = []
    parse_warnings: list[str] = []
    parse_notes: list[str] = []
    row_store_files = 0

    if not skip_xlsx_package_check:
        package_check = check_xlsx_package(file_path)
        for warning in package_check.get("warnings", []) or []:
            parse_warnings.append(f"[{file_path.name}] 包结构预检：{warning}")
        for note in package_check.get("notes", []) or []:
            parse_notes.append(f"[{file_path.name}] 包结构预检：{note}")

    try:
        formula_cell_count = count_formula_cells(file_path)
    except ImportError:
        return sheet_entries, ["缺少 openpyxl，无法解析 .xlsx/.xlsm 文件"], parse_notes, row_store_files
    except Exception as exc:  # noqa: BLE001
        formula_cell_count = 0
        parse_notes.append(f"[{file_path.name}] 公式统计失败，跳过重算：{exc}")

    if formula_cell_count > 0:
        recalc_result = recalc_excel_with_libreoffice(file_path, timeout_seconds=30)
        runtime_notes = recalc_result.get("runtime_notes", [])
        if isinstance(runtime_notes, list):
            for runtime_note in runtime_notes:
                parse_notes.append(f"[{file_path.name}] 公式重算运行环境：{runtime_note}")
        recalc_error = str(recalc_result.get("error", "")).strip()
        if recalc_error:
            parse_warnings.append(
                f"[{file_path.name}] 检测到 {formula_cell_count} 个公式单元格，但公式重算失败：{recalc_error}"
            )
        else:
            recalc_status = str(recalc_result.get("status", "")).strip().lower()
            total_errors = int(recalc_result.get("total_errors", 0) or 0)
            if recalc_status == "errors_found" and total_errors > 0:
                error_summary = recalc_result.get("error_summary", {})
                summary_parts: list[str] = []
                if isinstance(error_summary, dict):
                    for err_type, detail in error_summary.items():
                        if isinstance(detail, dict):
                            count_text = int(detail.get("count", 0) or 0)
                            summary_parts.append(f"{err_type}:{count_text}")
                summary_text = f"（{', '.join(summary_parts[:5])}）" if summary_parts else ""
                parse_warnings.append(
                    f"[{file_path.name}] 公式重算后发现 {total_errors} 个 Excel 错误{summary_text}，请修复后重试。"
                )
            else:
                parse_notes.append(
                    f"[{file_path.name}] 检测到 {formula_cell_count} 个公式单元格，已按内置 LibreOffice 主链路完成重算。"
                )

    unknown_extension_hits = 0
    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always")
        try:
            wb = open_workbook(file_path, data_only=True)
        except ImportError:
            return sheet_entries, ["缺少 openpyxl，无法解析 .xlsx/.xlsm 文件"], parse_notes, row_store_files
        except Exception as exc:  # noqa: BLE001
            parse_warnings.append(f"openpyxl 无法打开文件：{exc}")
            return sheet_entries, parse_warnings, parse_notes, row_store_files

        region_map = load_sheet_region_map(file_path, data_only=True) if getattr(wb, "read_only", False) else {}
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                min_row, max_row, min_col, max_col, region_note = region_map.get(sheet_name) or select_sheet_region(ws)
                if region_note:
                    parse_notes.append(f"[{file_path.name}/{sheet_name}] {region_note}")

                first_row = next(
                    ws.iter_rows(
                        min_row=min_row,
                        max_row=min_row,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True,
                    ),
                    tuple(),
                )
                headers = extract_headers(first_row)
                dup_warning = detect_duplicate_headers(headers)
                if dup_warning:
                    parse_warnings.append(f"[{file_path.name}/{sheet_name}] {dup_warning}")

                data_start_row = min_row + 1
                projected_headers = projected_headers_for_sheet(
                    projection_plan,
                    file_name=file_path.name,
                    file_path=file_path.as_posix(),
                    sheet=sheet_name,
                    headers=headers,
                )
                store_path_text = ""
                if not (projection_plan and not projected_headers):
                    store_path = rows_store_path(out_dir, file_path, sheet_name)
                    row_count = write_rows_store_stream(
                        rows_iter=(
                            (
                                row_idx,
                                row_to_map(
                                    headers,
                                    [json_friendly(v) for v in row],
                                ),
                            )
                            for row_idx, row in enumerate(
                                ws.iter_rows(
                                    min_row=data_start_row,
                                    max_row=max_row,
                                    min_col=min_col,
                                    max_col=max_col,
                                    values_only=True,
                                ),
                                start=data_start_row,
                            )
                        ),
                        projected_headers=projected_headers,
                        output_path=store_path,
                        chunk_size=row_chunk_size,
                    )
                    store_path_text = store_path.as_posix()
                    row_store_files += 1
                else:
                    row_count = max(0, int(max_row) - int(min_row))

                sheet_entries.append(
                    {
                        "sheet": sheet_name,
                        "headers": headers,
                        "projected_headers": projected_headers,
                        "row_count_estimate": row_count,
                        "rows_file": store_path_text,
                    }
                )
        finally:
            wb.close()

    for w in caught:
        msg = str(w.message)
        if "Unknown extension is not supported and will be removed" in msg:
            unknown_extension_hits += 1
            continue
        parse_warnings.append(f"openpyxl 告警：{msg}")

    if unknown_extension_hits > 0:
        xml_hits = scan_xlsx_extlst_sheet_xml(file_path)
        xml_text = f"；涉及: {', '.join(xml_hits[:6])}" if xml_hits else ""
        parse_notes.append(
            "检测到 openpyxl 不支持的工作表扩展（extLst）"
            f"，命中 {unknown_extension_hits} 次。根因通常是 Excel 高级特性扩展节点（如部分条件格式/数据验证/切片器）"
            f"{xml_text}。该扩展会被 openpyxl 忽略，但常规单元格值读取仍可继续。"
        )

    return sheet_entries, parse_warnings, parse_notes, row_store_files


def build_manifest(
    inputs: Path,
    out_dir: Path,
    run_id: str,
    fail_on_parser_warning: bool = True,
    compiled_rules_path: Path | None = None,
    row_chunk_size: int = 2000,
    skip_xlsx_package_check: bool = False,
) -> tuple[Path, str]:
    files = discover_input_files(inputs)
    if not files:
        raise ValueError(f"在 {inputs} 未发现可支持的输入文件")

    all_files: list[dict[str, Any]] = []
    input_hasher = hashlib.sha256()
    parser_warning_count = 0
    parser_note_count = 0
    projection_plan = build_projection_plan(compiled_rules_path)
    row_store_files = 0

    for file_path in files:
        digest = file_sha256(file_path)
        input_hasher.update(digest.encode("utf-8"))
        ext = file_path.suffix.lower()
        if ext in {".csv", ".tsv"}:
            sheet_entries, warnings, notes, delta_row_store = _ingest_text_file(
                file_path=file_path,
                out_dir=out_dir,
                projection_plan=projection_plan,
                row_chunk_size=row_chunk_size,
            )
        elif ext in {".xlsx", ".xlsm"}:
            sheet_entries, warnings, notes, delta_row_store = _ingest_xlsx_file(
                file_path=file_path,
                out_dir=out_dir,
                projection_plan=projection_plan,
                row_chunk_size=row_chunk_size,
                skip_xlsx_package_check=skip_xlsx_package_check,
            )
        else:
            sheet_entries = []
            warnings = [f"不支持的文件后缀：{ext}"]
            notes = []
            delta_row_store = 0

        parser_warning_count += len(warnings)
        parser_note_count += len(notes)
        row_store_files += delta_row_store

        all_files.append(
            {
                "path": file_path.as_posix(),
                "name": file_path.name,
                "extension": file_path.suffix.lower(),
                "size_bytes": file_path.stat().st_size,
                "sha256": digest,
                "sheets": sheet_entries,
                "parse_warnings": warnings,
                "parse_notes": notes,
            }
        )

    manifest = {
        "run_id": run_id,
        "generated_at": utc_now_iso(),
        "input_root": inputs.as_posix(),
        "files": all_files,
        "totals": {
            "file_count": len(all_files),
            "sheet_count": sum(len(x["sheets"]) for x in all_files),
            "parser_warning_count": parser_warning_count,
            "parser_note_count": parser_note_count,
            "row_store_file_count": row_store_files,
        },
        "input_hash": input_hasher.hexdigest(),
        "parse_options": {
            "excel_read_mode": "data_only",
            "formula_recalc": "builtin-libreoffice-recalc",
            "xlsx_package_check_enabled": not bool(skip_xlsx_package_check),
            "xlsx_package_check_policy": "warning",
            "row_chunk_size": int(row_chunk_size),
            "projection_enabled": bool(projection_plan),
            "projection_dataset_count": len(projection_plan),
        },
    }

    manifest_path = out_dir / "ingest_manifest.json"
    atomic_write_json(manifest_path, manifest)

    if fail_on_parser_warning and parser_warning_count > 0:
        raise ValueError("解析阶段产生告警，且 fail_on_parser_warning=true，已停止")

    return manifest_path, manifest["input_hash"]


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="扫描并解析 Excel/CSV/TSV 输入文件。")
    parser.add_argument("--inputs", required=True, help="输入文件或目录")
    parser.add_argument("--out", required=True, help="输出目录")
    parser.add_argument("--run-id", required=True, help="本次运行 ID")
    parser.add_argument(
        "--fail-on-parser-warning",
        dest="fail_on_parser_warning",
        action="store_true",
        help="若存在解析告警则直接失败（默认开启）",
    )
    parser.add_argument(
        "--allow-parser-warning",
        dest="fail_on_parser_warning",
        action="store_false",
        help="允许解析告警，不中断流程",
    )
    parser.add_argument(
        "--compiled-rules",
        default=None,
        help="可选：compiled_rules.json 路径，用于按规则列投影",
    )
    parser.add_argument(
        "--chunk-size",
        type=int,
        default=2000,
        help="行数据分块写入大小（默认 2000）",
    )
    parser.add_argument(
        "--skip-xlsx-package-check",
        action="store_true",
        help="跳过 xlsx/xlsm 包结构预检",
    )
    parser.set_defaults(fail_on_parser_warning=True)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    inputs = Path(args.inputs).resolve()
    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        manifest_path, _ = build_manifest(
            inputs=inputs,
            out_dir=out_dir,
            run_id=args.run_id,
            fail_on_parser_warning=args.fail_on_parser_warning,
            compiled_rules_path=Path(args.compiled_rules).resolve() if args.compiled_rules else None,
            row_chunk_size=max(1, int(args.chunk_size or 1)),
            skip_xlsx_package_check=bool(args.skip_xlsx_package_check),
        )
    except Exception as exc:  # noqa: BLE001
        print(f"[错误] parse_excel 执行失败：{exc}")
        return 1

    print(f"[成功] 输入清单输出：{manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
