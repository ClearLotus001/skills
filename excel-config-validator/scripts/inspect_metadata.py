"""轻量 Excel/CSV 元数据探查 — 仅输出 sheet 名与列头，不读取行数据。

用法：python scripts/inspect_metadata.py <file1> [file2 ...] [--out <output.json>]
输出：每个文件的 sheet 列表与各 sheet 的列名（JSON 格式）
      指定 --out 时写入文件；否则输出到 stdout。
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

# 确保 scripts/ 目录在导入路径中
sys.path.insert(0, str(Path(__file__).resolve().parent))


def inspect_xlsx(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h) if h is not None else "" for h in row]
            break
        sheets[name] = {"headers": headers}
    wb.close()
    return {"file": path.name, "type": "xlsx", "sheets": sheets}


def inspect_csv(path: Path) -> dict:
    import csv

    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
    sheet_name = path.stem
    return {"file": path.name, "type": "csv", "sheets": {sheet_name: {"headers": headers}}}


def inspect_xls(path: Path) -> dict:
    """使用 xlrd 探查 .xls 文件元数据。"""
    try:
        import xlrd  # type: ignore
    except ImportError:
        return {"file": path.name, "type": "xls", "error": "缺少 xlrd，无法解析 .xls"}

    book = xlrd.open_workbook(path.as_posix(), on_demand=True)
    try:
        sheets = {}
        for name in book.sheet_names():
            sh = book.sheet_by_name(name)
            headers = []
            if sh.nrows > 0:
                headers = [str(sh.cell_value(0, c)).strip() for c in range(sh.ncols)]
            sheets[name] = {"headers": headers}
    finally:
        book.release_resources()
    return {"file": path.name, "type": "xls", "sheets": sheets}


def inspect_xlsb(path: Path) -> dict:
    """使用 pyxlsb 探查 .xlsb 文件元数据。"""
    try:
        from pyxlsb import open_workbook  # type: ignore
    except ImportError:
        return {"file": path.name, "type": "xlsb", "error": "缺少 pyxlsb，无法解析 .xlsb"}

    with open_workbook(path.as_posix()) as wb:
        sheets = {}
        for name in wb.sheets:
            headers = []
            with wb.get_sheet(name) as sh:
                for row in sh.rows():
                    headers = [str(c.v).strip() if c.v is not None else "" for c in row]
                    break
            sheets[name] = {"headers": headers}
    return {"file": path.name, "type": "xlsb", "sheets": sheets}


def inspect_file(path: Path) -> dict:
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm"}:
        return inspect_xlsx(path)
    if ext == ".xls":
        return inspect_xls(path)
    if ext == ".xlsb":
        return inspect_xlsb(path)
    if ext == ".csv":
        return inspect_csv(path)
    return {"file": path.name, "type": "unknown", "error": f"unsupported extension: {ext}"}


def main() -> int:
    parser = argparse.ArgumentParser(description="探查 Excel/CSV 文件的 sheet 与列头元数据")
    parser.add_argument("files", nargs="+", help="要探查的文件路径")
    parser.add_argument("--out", default=None, help="可选：输出 JSON 文件路径（不指定则输出到 stdout）")
    args = parser.parse_args()

    results = []
    for f in args.files:
        p = Path(f).resolve()
        if not p.is_file():
            results.append({"file": str(f), "error": "file not found"})
            continue
        try:
            results.append(inspect_file(p))
        except Exception as exc:
            results.append({"file": p.name, "error": str(exc)})

    json_text = json.dumps(results, ensure_ascii=False, indent=2)

    if args.out:
        out_path = Path(args.out).resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json_text + "\n", encoding="utf-8")
    else:
        sys.stdout.write(json_text + "\n")
        sys.stdout.flush()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
